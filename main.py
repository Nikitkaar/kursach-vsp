from New_class import Shpala
from veroyatnost_raspredel_temp import GrafVerTemp
from progibi import Grafics_progibov
from temperature_diagramm import TempDiagramm
from izgibi import Izgibi
from initial_data import PodvizhnoySostav
from openpyxl import Workbook
from tkinter import messagebox
import tkinter as tk
import origin
from tkinter import ttk

class App:
    def __init__(self, root):

        print("Инициализация началась")
        self.root = root
        self.root.title("Ввод данных")

        # Создание виджета Canvas для прокрутки
        self.canvas = tk.Canvas(root)
        self.scroll_y = ttk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scroll_y.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scroll_y.pack(side="right", fill="y")

        # Создание полей ввода
        self.entries = {}
        entry_labels = [
            "Название станции",
            "t_min_min",
            "t_max_max",
            "redaction",
            "rail_type",
            "скорость",
            "capacity",
            "material_of_sleepers",
            "h",
            "t_opt",
            "val",
            "Type (элемент 1)",
            "Type (элемент 2)",
            "U[1]_лето_прямая",
            "U[2]_лето_кривая",
            "U[3]_зима_прямая",
            "U[4]_зима_кривая",
            "f[1]",
            "f[2]",
            "f[3]",
            "f[4]",
            "curve",
            "kd_locomotiva",
            "P_ct_vagona"
        ]

        # Добавление полей ввода в три столбца
        for index, label in enumerate(entry_labels):
            ttk.Label(self.scrollable_frame, text=label + ":").grid(row=index, column=0, sticky="w", pady=5)
            entry = ttk.Entry(self.scrollable_frame)
            entry.grid(row=index, column=1, pady=5)
            entry.bind('<Return>', self.focus_next)  # Привязка клавиши Enter к функции
            self.entries[label] = entry  # Сохраняем ссылки на поля ввода

        # Кнопка для копирования значений
        self.copy_button = ttk.Button(self.scrollable_frame, text="Скопировать", command=self.copy_to_clipboard)
        self.copy_button.grid(row=len(entry_labels) + 1, column=0, columnspan=2, pady=5)

        # Кнопка для вставки значений
        self.paste_button = ttk.Button(self.scrollable_frame, text="Вставить", command=self.paste_from_clipboard)
        self.paste_button.grid(row=len(entry_labels) + 2, column=0, columnspan=2, pady=5)

        # Кнопка для подтверждения ввода
        self.confirm_button = ttk.Button(self.scrollable_frame, text="Подтвердить", command=self.confirm)
        self.confirm_button.grid(row=len(entry_labels), column=0, columnspan=2, pady=20)

        print("Инициализация закончилась?")
    print("допустим инициализация на уровне аттрибутов класса")
    def focus_next(self, event):
        """Переход к следующему полю ввода"""
        current_widget = event.widget
        current_widget_index = list(self.entries.values()).index(current_widget)

        # Переход к следующему полю ввода
        next_widget = list(self.entries.values())[(current_widget_index + 1) % len(self.entries)]
        next_widget.focus_set()

    def copy_to_clipboard(self):
        """Копирует значения из полей ввода в буфер обмена"""
        clipboard_data = "\n".join([f"{label}: {entry.get()}" for label, entry in self.entries.items()])
        self.root.clipboard_clear()
        self.root.clipboard_append(clipboard_data)
        messagebox.showinfo("Копирование", "Данные скопированы в буфер обмена.")

    def paste_from_clipboard(self):
        """Вставляет значения из буфера обмена в поля ввода
        :param self:
        """
        clipboard_data = self.root.clipboard_get().strip().split('\n')
        for line, label in zip(clipboard_data, self.entries.keys()):
            try:
                key, value = line.split(':', 1)
                if key.strip() == label:
                    self.entries[label].delete(0, tk.END)  # Очистить текущее значение
                    self.entries[label].insert(0, value.strip())  # Вставить новое значение
            except ValueError:
                continue  # Игнорируем строки, которые не в формате "label: value"

    def confirm(self):
        print("Инициализация давно закончилась")
        # Считываем значения из полей ввода в глобальный словарь
        data = {}
        for field, entry in self.entries.items():
            value = entry.get().strip()  # Убираем пробелы

            if field in ["t_min_min", "t_max_max", "скорость", "h", "t_opt", "val", "capacity", "P_ct_vagona",
                         "curve",
                         "kd_locomotiva", "R"]:
                # Преобразуем числовые значения
                try:
                    if field in ["capacity", "kd_locomotiva"]:
                        value = float(value)
                    else:
                        value = int(value)
                except ValueError:
                    messagebox.showwarning("Ошибка", f"Введите корректное числовое значение для {field}.")
                    return

            data[field] = value  # Сохраняем значения

        # Формирование итоговых переменных
        self.station = data["Название станции"]
        self.val = data["val"]
        self.Type = [data["Type (элемент 1)"], data["Type (элемент 2)"]]
        self.U = [origin.characteristiki_puti[self.val - 1]["U"], origin.characteristiki_puti[self.val]["U"],
                  float(data["U[3]_зима_прямая"]), float(data["U[4]_зима_кривая"])]
        self.f = [float(data["f[1]"]), float(data["f[2]"]), float(data["f[3]"]), float(data["f[4]"])]
        self.capacity = data["capacity"]
        self.t_min_min = data["t_min_min"]
        self.t_max_max = data["t_max_max"]
        self.Ta = self.t_max_max + abs(self.t_min_min)
        self.kd_locomotiva = data["kd_locomotiva"]
        self.redaction = data["redaction"]
        self.v = data["скорость"]
        self.rail_type = data['rail_type']
        self.material_of_sleepers = data["material_of_sleepers"]
        self.P_ct_vagona = data["P_ct_vagona"]
        self.R = data["curve"]
        self.h = data["h"]
        self.t_opt = data["t_opt"]


        CriteriiProchnostiPuti_Vagon = {
            '>50': {"bkr": 1500, "bsh": 11, "bb": 2.6, "bz": 0.8},
            '50-25': {"bkr": 1600, "bsh": 15, "bb": 3, "bz": 0.8},
            '24-10': {"bkr": 2000, "bsh": 18, "bb": 3.5, "bz": 0.9},
            '<10': {"bkr": 3000, "bsh": 27, "bb": 4, "bz": 1.0}
        }
        # Получение критерия для capacity
        if self.capacity > 50:
            criteria = ">50"
        elif 50 >= self.capacity > 25:
            criteria = "50-25"
        elif 25 >= self.capacity > 10:
            criteria = "24-10"
        else:
            criteria = "<10"

        # Получение значения из пересечения строки [бк] и столбца с критерием
        бкр_vag = CriteriiProchnostiPuti_Vagon[f"{criteria}"]["bkr"]
        бш_vagona = CriteriiProchnostiPuti_Vagon[f"{criteria}"]["bsh"]
        бб_vagona = CriteriiProchnostiPuti_Vagon[f"{criteria}"]["bb"]
        бз_vagona = CriteriiProchnostiPuti_Vagon[f"{criteria}"]["bz"]

        CriteriiProchnostiPuti_Locomotiv = {
            ">50": {"bkr": 1900, "bsh": 12, "bb": 4, "bz": 1},
            "50-25": {"bkr": 2000, "bsh": 16, "bb": 4.2, "bz": 1},
            "24-10": {"bkr": 2400, "bsh": 20, "bb": 4.5, "bz": 1.1},
            "<10": {"bkr": 3400, "bsh": 30, "bb": 5, "bz": 1.2}
        }

        бкр_locomotiva = CriteriiProchnostiPuti_Locomotiv[criteria]["bkr"]
        бш_locomotiva = CriteriiProchnostiPuti_Locomotiv[criteria]["bsh"]
        бб_locomotiva = CriteriiProchnostiPuti_Locomotiv[criteria]["bb"]
        бз_locomotiva = CriteriiProchnostiPuti_Locomotiv[criteria]["bz"]

        Characteristici_locomotivov = {
            "ВЛ60": {'P_ct': 11500, 'q': 3080, 'G': 116, 'd': 125, 'n': 3, 'l_i': '230, 230, 0', 'fсг': 60},
            "ВЛ80": {'P_ct': 12000, 'q': 2760, 'G': 116, 'd': 125, 'n': 2, 'l_i': '300, 0, 0', 'fсг': 128},
            "ВЛ10": {'P_ct': 11500, 'q': 3160, 'G': 116, 'd': 125, 'n': 2, 'l_i': '300, 0, 0', 'fсг': 135},
            "ВЛ82": {'P_ct': 12500, 'q': 3160, 'G': 152, 'd': 125, 'n': 2, 'l_i': '200, 0, 0', 'fсг': 135},
            "ВЛ84": {'P_ct': 12500, 'q': 1730, 'G': 205, 'd': 125, 'n': 2, 'l_i': '285, 0, 0', 'fсг': 180},
            "ВЛ85": {'P_ct': 11500, 'q': 2760, 'G': 152, 'd': 125, 'n': 2, 'l_i': '290, 0, 0', 'fсг': 156},
            "ВЛ15": {'P_ct': 11500, 'q': 3050, 'G': 152, 'd': 125, 'n': 2, 'l_i': '290, 0, 0', 'fсг': 140},
            "ВЛ8": {'P_ct': 11500, 'q': 3170, 'G': 118, 'd': 120, 'n': 2, 'l_i': '320, 0, 0', 'fсг': 70},
            "ВЛ8М": {'P_ct': 11900, 'q': 3170, 'G': 118, 'd': 120, 'n': 2, 'l_i': '320, 0, 0', 'fсг': 70},
            "ВЛ22": {'P_ct': 11000, 'q': 3250, 'G': 135, 'd': 120, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 59},
            "ВЛ23": {'P_ct': 11500, 'q': 3170, 'G': 142, 'd': 120, 'n': 3, 'l_i': '220, 220, 0', 'fсг': 59},
            "ВЛ41": {'P_ct': 11500, 'q': 3075, 'G': 96, 'd': 120, 'n': 2, 'l_i': '210, 0, 0', 'fсг': 75},
            "ВЛ10у": {'P_ct': 12500, 'q': 3055, 'G': 152, 'd': 125, 'n': 2, 'l_i': '300, 0, 0', 'fсг': 155},
            "ВЛ12": {'P_ct': 12500, 'q': 3055, 'G': 152, 'd': 125, 'n': 2, 'l_i': '300, 0, 0', 'fсг': 155},
            "2ТЭ116": {'P_ct': 11500, 'q': 2230, 'G': 109, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 102},
            "2ТЭ10В": {'P_ct': 11500, 'q': 2230, 'G': 109, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 102},
            "2ТЭ10У": {'P_ct': 11500, 'q': 2230, 'G': 109, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 102},
            "ЗТЭ10У": {'P_ct': 11500, 'q': 2230, 'G': 109, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 102},
            "2ТЭ10М": {'P_ct': 11500, 'q': 2230, 'G': 109, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 102},
            "ЗТЭ10М": {'P_ct': 11500, 'q': 2230, 'G': 109, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 102},
            "ЗТЭ10В": {'P_ct': 11500, 'q': 2230, 'G': 109, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 102},
            "4ТЭ10С": {'P_ct': 11500, 'q': 2230, 'G': 109, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 102},
            "ТЭ3": {'P_ct': 10500, 'q': 2330, 'G': 14, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 57},
            "ЗТЭ3": {'P_ct': 10500, 'q': 2330, 'G': 14, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 57},
            "2ТЭ10УТ": {'P_ct': 11500, 'q': 2200, 'G': 112, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 102},
            "М62": {'P_ct': 9800, 'q': 2250, 'G': 129, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 69},
            "2М62": {'P_ct': 9800, 'q': 2250, 'G': 129, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 69},
            "М62У": {'P_ct': 10500, 'q': 2250, 'G': 110, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 104},
            "2М62У": {'P_ct': 10500, 'q': 2250, 'G': 110, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 104},
            "ЗМ62У": {'P_ct': 10500, 'q': 2250, 'G': 110, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 104},
            "ТЭМЗ": {'P_ct': 10600, 'q': 2980, 'G': 95, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 87},
            "ТЭП16": {'P_ct': 10600, 'q': 2980, 'G': 95, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 87},
            "ТЭМ17": {'P_ct': 10550, 'q': 2080, 'G': 110, 'd': 105, 'n': 3, 'l_i': '185, 185, 0', 'fсг': 104},
            "ТЭ10": {'P_ct': 11400, 'q': 2150, 'G': 129, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 69},
            "2ТЭ10": {'P_ct': 11400, 'q': 2150, 'G': 129, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 69},
            "2ТЭ10Л": {'P_ct': 10800, 'q': 2200, 'G': 119, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 69},
            "ТЭМ2": {'P_ct': 10000, 'q': 2330, 'G': 115, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 70},
            "TЭМ2УМ": {'P_ct': 10000, 'q': 2330, 'G': 115, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 70},
            "ТЭМ2УМТ": {'P_ct': 10000, 'q': 2330, 'G': 115, 'd': 105, 'n': 3, 'l_i': '210, 210, 0', 'fсг': 70},
            "ЧМЭЗ": {'P_ct': 10250, 'q': 2100, 'G': 166, 'd': 105, 'n': 3, 'l_i': '200, 200, 0', 'fсг': 49},
            "ВЛ-60": {'P_ct': 11500, 'q': 3080, 'G': 116, 'd': 125, 'n': 3, 'l_i': '230, 230, 0', 'fсг': 60}
        }

        P_ct = [Characteristici_locomotivov[self.Type[0]]['P_ct'], self.P_ct_vagona]
        q_loco = Characteristici_locomotivov[self.Type[0]]['q']
        G_loco = Characteristici_locomotivov[self.Type[0]]['G']
        d_locomotiva = Characteristici_locomotivov[self.Type[0]]['d']
        n_loco = Characteristici_locomotivov[self.Type[0]]['n']
        str_li_loc = Characteristici_locomotivov[self.Type[0]]['l_i']
        fcr_loc = Characteristici_locomotivov[self.Type[0]]['fсг']

        Characteristici_vagonov = {
            "4-осный": {'P_ct': 11000, 'q': 995, 'G': 200, 'd': 95, 'n': 2, 'l_i': '185, 0, 0', 'fсг': 48},
            "4-хосный": {'P_ct': 11000, 'q': 995, 'G': 200, 'd': 95, 'n': 2, 'l_i': '185, 0, 0', 'fсг': 48},
            "6-осный": {'P_ct': 10700, 'q': 1070, 'G': 195, 'd': 95, 'n': 3, 'l_i': '175, 175, 0', 'fсг': 50},
            "6-тиосный": {'P_ct': 10700, 'q': 1070, 'G': 195, 'd': 95, 'n': 3, 'l_i': '175, 175, 0', 'fсг': 50},
            "8-осный": {'P_ct': 10550, 'q': 995, 'G': 200, 'd': 95, 'n': 4, 'l_i': '185, 125, 185', 'fсг': 48},
            "8-миосный": {'P_ct': 10550, 'q': 995, 'G': 200, 'd': 95, 'n': 4, 'l_i': '185, 125, 185', 'fсг': 48}}

        q_vag = Characteristici_vagonov[self.Type[1]]['q']
        G_vagona = Characteristici_vagonov[self.Type[1]]['G']
        d_vagona = Characteristici_vagonov[self.Type[1]]['d']
        n_vag = Characteristici_vagonov[self.Type[1]]['n']
        str_li_vag = Characteristici_vagonov[self.Type[1]]['l_i']
        fcr_vag = Characteristici_vagonov[self.Type[1]]['fсг']

        print(str_li_loc)
        print(str_li_vag)
        # Разделить строку по запятым и преобразовать каждый элемент в числовой формат
        li_loc = [int(x) for x in str_li_loc.split(',')]
        li_vag = [int(x) for x in str_li_vag.split(',')]

        if self.redaction == "new":
            kd = [self.kd_locomotiva, round(0.1 + (0.2 * (self.v / fcr_vag)), 2)]  # ВВОД ИСХОДНЫХ ДАННЫХ
        else:  # and self.redaction != "old"
            kd = [0, 0]

        k_leto_curve = origin.characteristiki_puti[self.val]['k']
        k_leto_pryamo = origin.characteristiki_puti[self.val - 1]['k']
        L = origin.characteristiki_puti[self.val - 1]['L']
        alfa1 = origin.characteristiki_puti[self.val - 1]['α1']
        epsiland = origin.characteristiki_puti[self.val - 1]['eps']
        betaa = origin.characteristiki_puti[self.val - 1]['beta']
        gamma = origin.characteristiki_puti[self.val - 1]['gamma']

        W = origin.characteristiki_puti[self.val]['W(6)']
        alpha0 = origin.characteristiki_puti[self.val]['α0']
        omega = origin.characteristiki_puti[self.val]['ω']
        omega_a = origin.characteristiki_puti[self.val]['Ωа']
        b = origin.characteristiki_puti[self.val]['b']
        ae = origin.characteristiki_puti[self.val]['æ']

        # ВВОЖУ ИСХОДНЫЕ ДАННЫЕ
        # Локомотив/Вагон;
        q = [q_loco, q_vag]
        JestcostRessor = [G_loco, G_vagona]
        d = [d_locomotiva, d_vagona]
        n = [n_loco, n_vag]
        l_i = [li_loc, li_vag]
        print(l_i)
        curve = ["Прямая", self.R]  # РУЧНОЙ ВВОД ИЗ ДАНО
        # Прямая, Лето\Зима:     Кривая, Лето\Зима:
        k = [k_leto_curve, 0.0, k_leto_pryamo, 0.0]

        if self.rail_type == 'P50' or self.rail_type == 'Р50':
            J0 = 2018
        elif self.rail_type == 'P65' or self.rail_type == 'Р65':
            J0 = 3547
        else:
            J0 = 4490

        k[1] = round((self.U[1] / (4 * 2.1 * 10 ** 6 * J0)) ** 0.25, 5)
        k[3] = round((self.U[3] / (4 * 2.1 * 10 ** 6 * J0)) ** 0.25, 5)

        f0 = origin.wtf[curve[0]][self.Type[0]]
        f1 = origin.wtf[curve[1]][self.Type[0]]
        f2 = origin.wtf[curve[0]][self.Type[1]]
        f3 = origin.wtf[curve[1]][self.Type[1]]
        # Локомотив:     Вагон:
        # Прямая/Кривая; Прямая/Кривая
        if self.redaction == "new" or self.redaction == "новая" or self.redaction == "нов":  # при новой редакции ребята обычно сами дают значения f
            f = self.f
            if self.f == [0, 0, 0, 0]:
                f = [f0, f1, f2, f3]
        else:
            f = [f0, f1, f2, f3]

        sostavs = [
            PodvizhnoySostav(type_sostav=self.Type[0], season="Лето", curve=curve[0], q=q[0],
                             JestcostRessor=JestcostRessor[0],
                             d=d[0], n=n[0], l_i=l_i[0], f=f[0], e=0.047, u=self.U[0], k=k[2], P_ct=P_ct[0],
                             l_sh=55,
                             v=self.v, L=L,
                             alpha0=alpha0, W=W, rail_type=self.rail_type,
                             material_of_sleepers=self.material_of_sleepers,
                             omega=omega,
                             omega_a=omega_a, b=b, ae=ae, h=self.h, fcr=fcr_loc, kd=kd[0],
                             redaction=self.redaction),
            PodvizhnoySostav(type_sostav=self.Type[0], season="Зима", curve=curve[0], q=q[0],
                             JestcostRessor=JestcostRessor[0],
                             d=d[0], n=n[0], l_i=l_i[0], f=f[0], e=0.047, u=self.U[2], k=k[0], P_ct=P_ct[0],
                             l_sh=55,
                             v=self.v, L=L,
                             alpha0=alpha0, W=W, rail_type=self.rail_type,
                             material_of_sleepers=self.material_of_sleepers,
                             omega=omega,
                             omega_a=omega_a, b=b, ae=ae, h=self.h, fcr=fcr_loc, kd=kd[0],
                             redaction=self.redaction),
            PodvizhnoySostav(type_sostav=self.Type[0], season="Лето", curve=curve[1], q=q[0],
                             JestcostRessor=JestcostRessor[0],
                             d=d[0], n=n[0], l_i=l_i[0], f=f[1], e=0.047, u=self.U[1], k=k[1], P_ct=P_ct[0],
                             l_sh=51,
                             v=self.v, L=L,
                             alpha0=alpha0, W=W, rail_type=self.rail_type,
                             material_of_sleepers=self.material_of_sleepers,
                             omega=omega, omega_a=omega_a, b=b, ae=ae, h=self.h, fcr=fcr_loc, kd=kd[0],
                             redaction=self.redaction),
            PodvizhnoySostav(type_sostav=self.Type[0], season="Зима", curve=curve[1], q=q[0],
                             JestcostRessor=JestcostRessor[0],
                             d=d[0], n=n[0], l_i=l_i[0], f=f[1], e=0.047, u=self.U[3], k=k[3], P_ct=P_ct[0],
                             l_sh=51,
                             v=self.v, L=L,
                             alpha0=alpha0, W=W, rail_type=self.rail_type,
                             material_of_sleepers=self.material_of_sleepers,
                             omega=omega, omega_a=omega_a, b=b, ae=ae, h=self.h, fcr=fcr_loc, kd=kd[0],
                             redaction=self.redaction),
            PodvizhnoySostav(type_sostav=self.Type[1], season="Лето", curve=curve[0], q=q[1],
                             JestcostRessor=JestcostRessor[1],
                             d=d[1], n=n[1], l_i=l_i[1], f=f[2], e=0.067, u=self.U[0], k=k[2], P_ct=P_ct[1],
                             l_sh=55,
                             v=self.v, L=L,
                             alpha0=alpha0, W=W, rail_type=self.rail_type,
                             material_of_sleepers=self.material_of_sleepers,
                             omega=omega,
                             omega_a=omega_a, b=b, ae=ae, h=self.h, fcr=fcr_vag, kd=kd[1],
                             redaction=self.redaction),
            PodvizhnoySostav(type_sostav=self.Type[1], season="Зима", curve=curve[0], q=q[1],
                             JestcostRessor=JestcostRessor[1],
                             d=d[1], n=n[1], l_i=l_i[1], f=f[2], e=0.067, u=self.U[2], k=k[0], P_ct=P_ct[1],
                             l_sh=55,
                             v=self.v, L=L,
                             alpha0=alpha0, W=W, rail_type=self.rail_type,
                             material_of_sleepers=self.material_of_sleepers,
                             omega=omega,
                             omega_a=omega_a, b=b, ae=ae, h=self.h, fcr=fcr_vag, kd=kd[1],
                             redaction=self.redaction),
            PodvizhnoySostav(type_sostav=self.Type[1], season="Лето", curve=curve[1], q=q[1],
                             JestcostRessor=JestcostRessor[1],
                             d=d[1], n=n[1], l_i=l_i[1], f=f[3], e=0.067, u=self.U[1], k=k[1], P_ct=P_ct[1],
                             l_sh=51,
                             v=self.v, L=L,
                             alpha0=alpha0, W=W, rail_type=self.rail_type,
                             material_of_sleepers=self.material_of_sleepers,
                             omega=omega,
                             omega_a=omega_a, b=b, ae=ae, h=self.h, fcr=fcr_vag, kd=kd[1],
                             redaction=self.redaction),
            PodvizhnoySostav(type_sostav=self.Type[1], season="Зима", curve=curve[1], q=q[1],
                             JestcostRessor=JestcostRessor[1],
                             d=d[1], n=n[1], l_i=l_i[1], f=f[3], e=0.067, u=self.U[3], k=k[3], P_ct=P_ct[1],
                             l_sh=51,
                             v=self.v, L=L,
                             alpha0=alpha0, W=W, rail_type=self.rail_type,
                             material_of_sleepers=self.material_of_sleepers,
                             omega=omega,
                             omega_a=omega_a, b=b, ae=ae, h=self.h, fcr=fcr_vag, kd=kd[1],
                             redaction=self.redaction),
        ]

        """Минимальные температуры на прямой и кривой соответственно"""
        delta_t_p0_min = min(sostavs[1].delta_t_p(), sostavs[5].delta_t_p())
        delta_t_p1_min = min(sostavs[3].delta_t_p(), sostavs[7].delta_t_p())

        def indexxx_0():
            """возвращаем индекс sostavs, с минимальным delta_t_p на прямой"""
            global sostav
            for index, sostav in enumerate(sostavs):
                if sostav.delta_t_p() == delta_t_p0_min:
                    return index

        def indexxx_1():
            """возвращаем индекс sostavs, с минимальным delta_t_p на кривой"""
            global sostav
            for index, sostav in enumerate(sostavs):
                if sostav.delta_t_p() == delta_t_p1_min:
                    return index

        if self.rail_type == 'P50':
            "Площадь поперечного сечения рельса"
            F = 65.99
        else:
            F = 82.56

        # РАБОТА С ЭКСЕЛЬ, ЗАПИСЬ ДАННЫХ В ТАБЛИЦУ (с помощью специальной вставки они переходят в ворд-документ отчета)
        # ГЛАВНАЯ ТАБЛИЦА

        workbook_3 = Workbook()
        sheet = workbook_3.active

        for i, sostav in enumerate(sostavs, start=1):

            sheet.cell(row=1, column=4, value=self.Type[0])
            sheet.cell(row=1, column=5, value=self.Type[1])

            sheet.cell(row=2, column=i, value=sostav.curve)

            sheet.cell(row=3, column=i, value=sostav.season)

            sheet.cell(row=4, column=4, value=P_ct[0])
            sheet.cell(row=4, column=5, value=P_ct[1])
            sheet.cell(row=4, column=6, value="Pcт")

            sheet.cell(row=5, column=1, value=self.v)
            sheet.cell(row=5, column=2, value="Скорость")

            sheet.cell(row=6, column=4, value=q[0])
            sheet.cell(row=6, column=5, value=q[1])
            sheet.cell(row=6, column=6, value="q")

            sheet.cell(row=7, column=4, value=JestcostRessor[0])
            sheet.cell(row=7, column=5, value=JestcostRessor[1])
            sheet.cell(row=7, column=6, value="JestcostRessor")

            sheet.cell(row=8, column=4, value=d[0])
            sheet.cell(row=8, column=5, value=d[1])
            sheet.cell(row=8, column=6, value="d")

            sheet.cell(row=9, column=4, value=n[0])
            sheet.cell(row=9, column=5, value=n[1])
            sheet.cell(row=9, column=i, value="n")

            sheet.cell(row=10, column=i, value=self.rail_type)

            sheet.cell(row=11, column=i, value=round(sostav.f, 2))
            sheet.cell(row=11, column=9, value="f")

            sheet.cell(row=12, column=4, value=sostavs[0].e)
            sheet.cell(row=12, column=5, value=sostavs[4].e)
            sheet.cell(row=12, column=6, value="e")

            sheet.cell(row=13, column=4, value=sostavs[0].z_max()[1])
            sheet.cell(row=13, column=5, value=sostavs[4].z_max()[1])
            sheet.cell(row=13, column=6, value="z_max")

            sheet.cell(row=133, column=i, value=sostav.Ekv_gruzi_η())
            sheet.cell(row=133, column=9, value="Ekv_gruzi_η")

            sheet.cell(row=134, column=i, value=sostav.Ekv_gruzi_µ())
            sheet.cell(row=134, column=9, value="Ekv_gruzi_µ")

            if sostav.n == 4:
                inf = f'{sostav.l_i[0]}+{sostav.l_i[1]}+{sostav.l_i[2]}'
            elif sostav.n == 3:
                inf = f'{sostav.l_i[0]}+{sostav.l_i[1]}'
            else:
                inf = str(sostav.l_i[0])
            sheet.cell(row=14, column=i, value=inf)

            sheet.cell(row=15, column=4, value=round(sostavs[0].z_max()[0], 2))
            sheet.cell(row=15, column=5, value=round(sostavs[4].z_max()[0], 2))
            sheet.cell(row=15, column=6, value="z_max")

            sheet.cell(row=16, column=i, value=self.material_of_sleepers)

            sheet.cell(row=17, column=i, value=sostav.u)
            sheet.cell(row=17, column=9, value="u")

            sheet.cell(row=18, column=i, value=sostav.k)
            sheet.cell(row=18, column=9, value="k")

            sheet.cell(row=19, column=1, value=L)
            sheet.cell(row=19, column=2, value="L")

            sheet.cell(row=20, column=i, value=sostav.W)

            sheet.cell(row=21, column=i, value=sostav.alpha0)

            sheet.cell(row=22, column=i, value=sostav.omega)

            sheet.cell(row=23, column=i, value=sostav.omega_a)

            sheet.cell(row=24, column=i, value=sostav.b)

            inf = sostav.p_max_p()
            sheet.cell(row=25, column=i, value=inf)
            sheet.cell(row=25, column=9, value='p_max_p')

            inf = sostav.p_cp_p()
            sheet.cell(row=26, column=i, value=inf)

            inf = sostav.p_cp()
            sheet.cell(row=27, column=i, value=inf)

            inf = sostav.s_p()
            sheet.cell(row=28, column=i, value=inf)

            inf = sostav.s_np()
            sheet.cell(row=29, column=i, value=inf)

            inf = sostav.s_nnk()
            sheet.cell(row=30, column=i, value=inf)

            inf = sostav.s_ink()
            sheet.cell(row=31, column=i, value=inf)

            inf = sostav.s()
            sheet.cell(row=32, column=i, value=inf)

            inf = sostav.p_max_ver()
            sheet.cell(row=33, column=i, value=inf)
            sheet.cell(row=33, column=9, value='p_max_ver')

            inf = sostav.pi_4k()
            sheet.cell(row=34, column=i, value=inf)

            inf = sostav.pi_5_4k()
            sheet.cell(row=35, column=i, value=inf)

            inf = sostav.pi_3_4k()
            sheet.cell(row=36, column=i, value=inf)

            inf = sostav.pi_7_4k()
            sheet.cell(row=37, column=i, value=inf)

            inf = round(sostav.P_I_ekv(), 0)
            sheet.cell(row=38, column=i, value=inf)

            inf = sostav.Muu2()
            sheet.cell(row=39, column=i, value=inf)

            inf = sostav.Muu3()
            sheet.cell(row=40, column=i, value=inf)

            inf = sostav.Muu4()
            sheet.cell(row=41, column=i, value=inf)

            inf = sostav.Sigma_Muu()
            sheet.cell(row=42, column=i, value=inf)

            inf = round(sostav.P_II_ekv())
            sheet.cell(row=43, column=i, value=inf)

            inf = sostav.N_2()
            sheet.cell(row=44, column=i, value=inf)

            inf = sostav.N_3()
            sheet.cell(row=45, column=i, value=inf)

            inf = sostav.N_4()
            sheet.cell(row=46, column=i, value=inf)

            inf = sostav.Sigma_N()
            sheet.cell(row=47, column=i, value=inf)

            inf = sostav.RaschetnayaOS_N()
            sheet.cell(row=48, column=i, value=inf)
            sheet.cell(row=48, column=9, value="номер оси n")

            inf = sostav.sigma_kp()
            sheet.cell(row=49, column=i, value=inf)

            inf = sostav.sigma_sh()
            sheet.cell(row=50, column=i, value=inf)

            inf = sostav.sigma_br()
            sheet.cell(row=51, column=i, value=inf)

            inf = sostav.ae
            sheet.cell(row=52, column=i, value=inf)

            sheet.cell(row=53, column=i, value=sostav.NNN(55))
            sheet.cell(row=53, column=9, value="Тета от длинны шпалы")

            sheet.cell(row=54, column=i, value=J0)
            sheet.cell(row=54, column=9, value="Момент инерции рельса")

            inf = sostav.m()
            sheet.cell(row=55, column=i, value=inf)

            inf = sostav.sigma_h2()
            sheet.cell(row=56, column=i, value=inf)
            sheet.cell(row=56, column=9, value="sigma_h2()")

            sheet.cell(row=57, column=i, value=sostav.A())
            sheet.cell(row=57, column=9, value='А, коэффициент расстояния между шпал')

            sheet.cell(row=58, column=i, value=sostav.C1())
            sheet.cell(row=58, column=9, value='C1')

            inf = sostav.l_i[0]
            sheet.cell(row=59, column=i, value=inf)
            sheet.cell(row=59, column=9, value='l_i[0]')

            inf = sostav.l_i[1]
            sheet.cell(row=60, column=i, value=inf)
            sheet.cell(row=60, column=9, value='l_i[1]')

            inf = sostav.l_i[2]
            sheet.cell(row=61, column=i, value=inf)
            sheet.cell(row=61, column=9, value='l_i[2]')

            sheet.cell(row=63, column=8, value=self.h)
            sheet.cell(row=63, column=9, value='высота насыпи')

            sheet.cell(row=62, column=i, value=бз_vagona)
            sheet.cell(row=62, column=9, value='[бз_Вагон]')

            inf = sostav.C2()
            sheet.cell(row=64, column=i, value=inf)
            sheet.cell(row=64, column=9, value='C2')

            inf = sostav.delta_t_p()
            sheet.cell(row=67, column=i, value=inf)
            sheet.cell(row=67, column=9, value="[∆t_р]")

            inf = sostav.k3
            sheet.cell(row=68, column=i, value=inf)

            inf = sostav.curve
            sheet.cell(row=70, column=i, value=inf)

            inf = sostav.xm()[0]
            sheet.cell(row=71, column=i, value=inf)

            inf = sostav.xm()[1]
            sheet.cell(row=72, column=i, value=inf)

            inf = sostav.xm()[2]
            sheet.cell(row=73, column=i, value=inf)

            inf = sostav.xm()[3]
            sheet.cell(row=74, column=i, value=inf)

            sheet.cell(row=75, column=1, value=бкр_locomotiva)
            sheet.cell(row=76, column=1, value=бш_locomotiva)
            sheet.cell(row=77, column=1, value=бб_locomotiva)
            sheet.cell(row=75, column=2, value="Loko_[бкр]")
            sheet.cell(row=76, column=2, value="Loko_[бш]")
            sheet.cell(row=77, column=2, value="Loko_[бб]")

            sheet.cell(row=78, column=1, value=бкр_vag)
            sheet.cell(row=79, column=1, value=бш_vagona)
            sheet.cell(row=80, column=1, value=бб_vagona)
            sheet.cell(row=78, column=2, value="Vag_[бкр]")
            sheet.cell(row=79, column=2, value="Vag_[бш]")
            sheet.cell(row=80, column=2, value="Vag_[бб]")

            inf = sostav.xn()[2]
            sheet.cell(row=81, column=i, value=inf)
            sheet.cell(row=81, column=9, value="xn")

            inf = round(sostav.summa1(), 5)
            sheet.cell(row=82, column=i, value=inf)
            sheet.cell(row=82, column=9, value="сигма тета 1 шпала")

            sheet.cell(row=83, column=i, value=fcr_loc)
            sheet.cell(row=83, column=9, value='fcr_loc')

            sheet.cell(row=84, column=i, value=fcr_vag)
            sheet.cell(row=84, column=9, value='fcr_vag')

            sheet.cell(row=85, column=i, value=alfa1)
            sheet.cell(row=85, column=9, value='alfa1')

            sheet.cell(row=86, column=i, value=betaa)
            sheet.cell(row=86, column=9, value='Beta')

            sheet.cell(row=87, column=i, value=gamma)
            sheet.cell(row=87, column=9, value='Gamma')

            sheet.cell(row=88, column=i, value=epsiland)
            sheet.cell(row=88, column=9, value='Epsiland')

            sheet.cell(row=89, column=i, value=kd[0])
            sheet.cell(row=89, column=9, value='kd_loc')

            sheet.cell(row=90, column=i, value=kd[1])
            sheet.cell(row=90, column=9, value='kd_vag')

            sheet.cell(row=96, column=i, value=sostav.p_max_p_zmax())
            sheet.cell(row=96, column=9, value='Pmax_p_zmax')

            sheet.cell(row=97, column=i, value=sostav.p_max_p_kd())
            sheet.cell(row=97, column=9, value='Pmax_p_kd')

            inf = round(sostav.summa2(), 5)
            sheet.cell(row=95, column=i, value=inf)
            sheet.cell(row=95, column=9, value="сигма тета 3 шпала")

            inf = sostav.P_II_ekvONE()
            sheet.cell(row=100, column=i, value=inf)
            sheet.cell(row=100, column=9, value="P_II_ekvONE")

            inf = sostav.P_II_ekvThree()
            sheet.cell(row=101, column=i, value=inf)
            sheet.cell(row=101, column=9, value="P_II_ekvThree")

            inf = sostavs[4].sigma_b1()
            sheet.cell(row=102, column=i, value=inf)
            sheet.cell(row=102, column=9, value='sigma_b1')

            inf = sostavs[4].sigma_h1()
            sheet.cell(row=103, column=i, value=inf)
            sheet.cell(row=103, column=9, value='sigma_h1')

            inf = sostavs[4].sigma_b3()
            sheet.cell(row=104, column=i, value=inf)
            sheet.cell(row=104, column=9, value='sigma_b3')

            inf = sostavs[4].sigma_h3()
            sheet.cell(row=105, column=i, value=inf)
            sheet.cell(row=105, column=9, value='sigma_h3')

            inf = sostavs[4].sigma_h()
            sheet.cell(row=106, column=i, value=inf)
            sheet.cell(row=106, column=9, value="∑_h")

            inf = sostav.AA()
            sheet.cell(row=107, column=i, value=inf)
            sheet.cell(row=107, column=9, value="Параметр А Першин")

            inf = sostav.AA1()
            sheet.cell(row=108, column=i, value=inf)
            sheet.cell(row=108, column=9, value="Параметр µ Першин")

            sheet.cell(row=113, column=i, value=delta_t_p0_min)
            sheet.cell(row=113, column=9, value="[∆t_р_min]Прямая")
            sheet.cell(row=114, column=i,
                       value=round(sostavs[indexxx_0()].sigma_kp() * 1.3 + 25.2 * delta_t_p0_min, 2))
            sheet.cell(row=114, column=9, value="нормальные суммарные напряжения(Прямая)")

            sheet.cell(row=115, column=i, value=delta_t_p1_min)
            sheet.cell(row=115, column=9, value="[∆t_р_min]Кривая")
            sheet.cell(row=116, column=i,
                       value=round(sostavs[indexxx_1()].sigma_kp() * 1.3 + 25.2 * delta_t_p1_min, 2))
            sheet.cell(row=116, column=9, value="нормальные суммарные напряжения(Кривая)")

            sheet.cell(row=135, column=i, value=sostavs[4].Ekv_gruzi_η_shpala_1())
            sheet.cell(row=135, column=9, value="Ekv_gruzi_η_shpala_1")

            sheet.cell(row=136, column=i, value=sostavs[4].Ekv_gruzi_η_shpala_3())
            sheet.cell(row=136, column=9, value="Ekv_gruzi_η_shpala_3")

            sheet.cell(row=117, column=i, value=25.2 * delta_t_p0_min)
            sheet.cell(row=118, column=i, value=25.2 * delta_t_p1_min)
            sheet.cell(row=117, column=9, value="σ_t(Прямая)")
            sheet.cell(row=118, column=9, value="σ_t(Кривая)")

            sigma_kp0 = sostavs[indexxx_0()].sigma_kp()
            sheet.cell(row=121, column=i, value=sigma_kp0)
            sheet.cell(row=121, column=9, value="σ_кп^зима_ПРямая")

            sigma_kp1 = sostavs[indexxx_1()].sigma_kp()
            sheet.cell(row=122, column=i, value=sigma_kp1)
            sheet.cell(row=122, column=9, value="σ_кп^зима_Кривая")

            AA0 = sostavs[indexxx_0()].AA()
            AA1 = sostavs[indexxx_0()].AA1()
            P_k0 = round((AA0 * 0.9 * 1 * 1.03) / (2 ** AA1), 2)
            sheet.cell(row=123, column=i, value=P_k0)
            sheet.cell(row=123, column=9, value="критическая температурная сила Pк(Прямая)")

            AA2 = sostavs[indexxx_1()].AA()
            AA3 = sostavs[indexxx_1()].AA1()
            P_k1 = round((AA2 * 0.9 * 1.08 * 1.05) / (3 ** AA3), 2)
            sheet.cell(row=124, column=i, value=P_k1)
            sheet.cell(row=124, column=9, value="критическая температурная сила Pк(Кривая)")

            kgs0 = round(P_k0 * 101.971621297793, 2)
            sheet.cell(row=125, column=i, value=kgs0)
            sheet.cell(row=125, column=9, value="Pк(Прямая)кгс")

            kgs1 = round(P_k1 * 101.971621297793, 2)
            sheet.cell(row=126, column=i, value=kgs1)
            sheet.cell(row=126, column=9, value="Pк(Кривая)кгс")

            P_norm0 = round(kgs0 / 1.5, 2)
            sheet.cell(row=127, column=i, value=P_norm0)
            sheet.cell(row=127, column=9, value="[P]Прямая")

            P_norm1 = round(kgs1 / 1.5, 2)
            sheet.cell(row=128, column=i, value=P_norm1)
            sheet.cell(row=128, column=9, value="[P]Кривая")

            sheet.cell(row=129, column=i, value=F)
            sheet.cell(row=129, column=9, value="F")

            sheet.cell(row=130, column=i, value=F * 2)
            sheet.cell(row=130, column=9, value="F * 2")

            t_y = round(P_norm0 / (25 * 2 * F))
            sheet.cell(row=131, column=i, value=t_y)
            sheet.cell(row=131, column=9, value="[∆t_уПрямая]")
            t_y_curve = round(P_norm1 / (25 * 2 * F))
            sheet.cell(row=132, column=i, value=t_y_curve)
            sheet.cell(row=132, column=9, value="[∆t_у_curve]")

            sheet.cell(row=141, column=i, value=delta_t_p0_min + round(P_norm0 / (25 * 2 * F), 2) - 10)
            sheet.cell(row=141, column=9, value="[T] прямая")
            sheet.cell(row=140, column=i, value=delta_t_p1_min + round(P_norm1 / (25 * 2 * F), 2) - 10)
            sheet.cell(row=140, column=9, value="[T] кривая")
            sheet.cell(row=139, column=i, value=self.Ta)
            sheet.cell(row=139, column=9, value="Tа")
            sheet.cell(row=138, column=i, value=self.t_min_min)
            sheet.cell(row=138, column=9, value="t_min_min")
            sheet.cell(row=137, column=i, value=self.t_max_max)
            sheet.cell(row=137, column=9, value="t_max_max")

            t_max_zakr = min(self.t_min_min + delta_t_p0_min, self.t_max_max)
            sheet.cell(row=142, column=i, value=t_max_zakr)
            sheet.cell(row=142, column=9, value="t_max_zakr")
            t_max_zakr_curve = min(self.t_min_min + delta_t_p1_min, self.t_max_max)
            sheet.cell(row=143, column=i, value=t_max_zakr_curve)
            sheet.cell(row=143, column=9, value="t_max_zakr_curve")
            t_min_zakr = max(self.t_max_max - t_y, self.t_min_min)
            sheet.cell(row=144, column=i, value=t_min_zakr)
            sheet.cell(row=144, column=9, value='t_min_zakr')
            t_min_zakr_curve = max(self.t_max_max - t_y_curve, self.t_min_min)
            sheet.cell(row=145, column=i, value=t_min_zakr_curve)
            sheet.cell(row=145, column=9, value='t_min_zakr_curve')
            sheet.cell(row=146, column=9, value='Грузонапряженность')
            sheet.cell(row=146, column=i, value=self.capacity)
            sheet.cell(row=147, column=9, value='Станция')
            sheet.cell(row=147, column=i, value=self.station)

            vvv = GrafVerTemp(t_max_max=self.t_max_max, t_min_min=self.t_min_min,
                              t_min_zakr=min(t_min_zakr, t_min_zakr_curve), t_max_zakr=t_max_zakr,
                              t_y=min(t_y, t_y_curve), delta_t_p=min(delta_t_p1_min, delta_t_p0_min),
                              t_opt=self.t_opt)
            vvv.grafic_maker_()

            aaa = TempDiagramm(t_max_max=self.t_max_max, t_min_min=self.t_min_min, t_y_curve=t_y_curve,
                               curve=curve[1],
                               t_min_zakr=t_min_zakr, t_max_zakr=t_max_zakr, t_min_zakr_curve=t_min_zakr_curve,
                               t_max_zakr_curve=t_max_zakr_curve, t_y=t_y, delta_t_p1_min=delta_t_p1_min,
                               delta_t_p0_min=delta_t_p0_min)
            aaa.grafic_maker()

        # Уже заняты 133 и 134
        # Почему-то возвращают только первое вхождение если строки кода стоят здесь, а не выше

        workbook_3.save('УЛЬТИМАТИВНАЯ_Формулы.xlsx')

        xxx = Shpala(sostavs=sostavs[4])
        xxx.grafic()  # Вызов метода для построения графиков

        yyy = Grafics_progibov(sostavs=sostavs, k=k, l_i_loc=li_loc, l_i_vag=li_vag)
        yyy.grafic_mader()

        zzz = Izgibi(sostavs=sostavs, k=k)
        zzz.grafic_mader()

        messagebox.showinfo("Успех", "Все данные успешно введены!")
        self.root.destroy()  # Закрыть окно после подтверждения
    # Запуск приложения
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()


