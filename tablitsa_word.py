
from openpyxl import Workbook

n1 = 3
n2 = 4
header = ["Условия эксплуатации", "№ оси", "xi", "kсм^-1",	"kx", "ɳi",	"∑ɳ", "PIIэкв",	"σщ", "σб"]
header_vert = ["Локомотив", "Вагон", "зима", "лето"]
tablitsa = Workbook()
sheet = tablitsa.active




sheet.cell(row=1, column=1, value="Условия эксплуатации")
sheet.cell(row=1, column=4, value="№ оси")
sheet.cell(row=1, column=5, value="xi")
sheet.cell(row=1, column=6, value="kсм^-1")
sheet.cell(row=1, column=7, value="kx")
sheet.cell(row=1, column=8, value="ɳi")
sheet.cell(row=1, column=9, value="∑ɳ")
sheet.cell(row=1, column=10, value="PIIэкв")
sheet.cell(row=1, column=11, value="σщ")
sheet.cell(row=1, column=12, value="σб")
sheet.cell(row=2, column=1, value="Локомотив")
sheet.cell(row=n1*4+2, column=1, value="Вагон")

sheet.merge_cells("A1:C1")
sheet.merge_cells(f"A2:A{n1*4+1}")
sheet.merge_cells(f"A{n1*4+2}:A{n1*4+1+n2*4}")
sheet.merge_cells(f"C2:C{n1+1}")
sheet.merge_cells(f"C{n1+2}:C{2*n1+1}")
sheet.merge_cells(f"C{2*n1+2}:C{3*n1+1}")
sheet.merge_cells(f"C{3*n1+2}:C{4*n1+1}")
sheet.merge_cells(f"C{n1*4+2}:C{n1*4+1+n2}")
sheet.merge_cells(f"C{n1*4+2+n2}:C{n1*4+1+2*n2}")
sheet.merge_cells(f"C{n1*4+2+2*n2}:C{n1*4+1+3*n2}")
sheet.merge_cells(f"C{n1*4+2+3*n2}:C{n1*4+1+4*n2}")
sheet.merge_cells(f"B2:B{2*n1+1}")
sheet.merge_cells(f"B{2*n1+2}:B{4*n1+1}")
sheet.merge_cells(f"B{4*n1+2}:B{4*n1+1+2*n2}")
sheet.merge_cells(f"B{4*n1+2+2*n2}:B{4*n1+1+4*n2}")

sheet.merge_cells(f"F2:F{n1+1}")
sheet.merge_cells(f"F{n1+2}:F{2*n1+1}")
sheet.merge_cells(f"F{2*n1+2}:F{3*n1+1}")
sheet.merge_cells(f"F{3*n1+2}:F{4*n1+1}")
sheet.merge_cells(f"F{n1*4+2}:F{n1*4+1+n2}")
sheet.merge_cells(f"F{n1*4+2+n2}:F{n1*4+1+2*n2}")
sheet.merge_cells(f"F{n1*4+2+2*n2}:F{n1*4+1+3*n2}")
sheet.merge_cells(f"F{n1*4+2+3*n2}:F{n1*4+1+4*n2}")

sheet.merge_cells(f"I2:I{n1+1}")
sheet.merge_cells(f"I{n1+2}:I{2*n1+1}")
sheet.merge_cells(f"I{2*n1+2}:I{3*n1+1}")
sheet.merge_cells(f"I{3*n1+2}:I{4*n1+1}")
sheet.merge_cells(f"I{n1*4+2}:I{n1*4+1+n2}")
sheet.merge_cells(f"I{n1*4+2+n2}:I{n1*4+1+2*n2}")
sheet.merge_cells(f"I{n1*4+2+2*n2}:I{n1*4+1+3*n2}")
sheet.merge_cells(f"I{n1*4+2+3*n2}:I{n1*4+1+4*n2}")

sheet.merge_cells(f"J2:J{n1+1}")
sheet.merge_cells(f"J{n1+2}:J{2*n1+1}")
sheet.merge_cells(f"J{2*n1+2}:J{3*n1+1}")
sheet.merge_cells(f"J{3*n1+2}:J{4*n1+1}")
sheet.merge_cells(f"J{n1*4+2}:J{n1*4+1+n2}")
sheet.merge_cells(f"J{n1*4+2+n2}:J{n1*4+1+2*n2}")
sheet.merge_cells(f"J{n1*4+2+2*n2}:J{n1*4+1+3*n2}")
sheet.merge_cells(f"J{n1*4+2+3*n2}:J{n1*4+1+4*n2}")

sheet.merge_cells(f"K2:K{n1+1}")
sheet.merge_cells(f"K{n1+2}:K{2*n1+1}")
sheet.merge_cells(f"K{2*n1+2}:K{3*n1+1}")
sheet.merge_cells(f"K{3*n1+2}:K{4*n1+1}")
sheet.merge_cells(f"K{n1*4+2}:K{n1*4+1+n2}")
sheet.merge_cells(f"K{n1*4+2+n2}:K{n1*4+1+2*n2}")
sheet.merge_cells(f"K{n1*4+2+2*n2}:K{n1*4+1+3*n2}")
sheet.merge_cells(f"K{n1*4+2+3*n2}:K{n1*4+1+4*n2}")

sheet.merge_cells(f"L2:L{n1+1}")
sheet.merge_cells(f"L{n1+2}:L{2*n1+1}")
sheet.merge_cells(f"L{2*n1+2}:L{3*n1+1}")
sheet.merge_cells(f"L{3*n1+2}:L{4*n1+1}")
sheet.merge_cells(f"L{n1*4+2}:L{n1*4+1+n2}")
sheet.merge_cells(f"L{n1*4+2+n2}:L{n1*4+1+2*n2}")
sheet.merge_cells(f"L{n1*4+2+2*n2}:L{n1*4+1+3*n2}")
sheet.merge_cells(f"L{n1*4+2+3*n2}:L{n1*4+1+4*n2}")

tablitsa.save('tablitsa.xlsx')
tablitsa.close()

